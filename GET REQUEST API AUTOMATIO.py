import pandas as pd
import requests
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Define the input and output file names
input_file = '/home/vignesh/PycharmProjects/Api_Testing_Automation/GET API DATA DRIVEN.xlsx'
output_file = 'output.xlsx'

# Define the headers to be sent with the GET request
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

# Read the input xlsx file into a pandas DataFrame
df = pd.read_excel(input_file)

# Add a new column for the response
df['Response'] = ""

# Send a GET request for each row in the input DataFrame
for i, row in df.iterrows():
    url = row['URL']
    # Send a GET request with the headers
    response = requests.get(url, headers=headers)
    # Get the response content
    content = response.content.decode('utf-8')
    # Add the response content to the row in the DataFrame
    df.at[i, 'Response'] = content

# Write the output DataFrame to an xlsx file with formatting
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'))
for row in ws[2:ws.max_row]:
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = Border(bottom=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'))
for col in ws.columns:
    column = col[0].column_letter
    ws.column_dimensions[column].width = max(df[col[0].value].astype(str).str.len().max(), len(col[0].value))
wb.save(output_file)
